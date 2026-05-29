import asyncio
import json
import os
import re
import sys
import tkinter as tk
from dataclasses import dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin

from openpyxl import Workbook, load_workbook
from playwright.async_api import BrowserContext, Page, Playwright, async_playwright


def get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    try:
        return Path.cwd().resolve()
    except Exception:
        return Path.cwd()


BASE_DIR = get_base_dir()

BASE_URL = "https://school.mos.ru"
JOURNALS_URL = f"{BASE_URL}/teacher/study-process/journal/my"
APP_TITLE = "Запуск обработки журналов МЭШ"


def get_local_appdata_dir() -> Path:
    raw = os.environ.get("LOCALAPPDATA")
    if raw:
        return Path(raw)
    return Path.home() / "AppData" / "Local"


def get_yandex_browser_candidates() -> List[Path]:
    local_appdata = get_local_appdata_dir()
    candidates = [
        local_appdata / "Yandex" / "YandexBrowser" / "Application" / "browser.exe",
        Path(os.environ.get("PROGRAMFILES", r"C:\Program Files")) / "Yandex" / "YandexBrowser" / "Application" / "browser.exe",
        Path(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)")) / "Yandex" / "YandexBrowser" / "Application" / "browser.exe",
    ]
    uniq: List[Path] = []
    seen = set()
    for item in candidates:
        key = str(item).lower()
        if key not in seen:
            uniq.append(item)
            seen.add(key)
    return uniq


def find_yandex_browser_path() -> Optional[Path]:
    for candidate in get_yandex_browser_candidates():
        if candidate.exists():
            return candidate
    return None


def get_default_user_data_dir() -> Path:
    return get_local_appdata_dir() / "Yandex" / "YandexBrowser" / "User Data"


def clean_path_from_ui(raw: str) -> Path:
    return Path(raw.strip().strip('"')).expanduser()


@dataclass
class AppConfig:
    yandex_browser_path: Path
    user_data_dir: Path
    profile_directory: str = "Default"
    xlsx_path: Path = BASE_DIR / "student_skills_recommendations.xlsx"
    headless: bool = False
    max_concurrent_journals: int = 1
    request_timeout: int = 60_000
    save_enable_timeout: int = 12_000

    # Дополнительные ожидания для медленного mos.ru
    journal_extra_stabilization_ms: int = 1_000
    before_student_click_ms: int = 1_000
    after_student_popup_visible_ms: int = 1_000
    after_side_page_visible_ms: int = 1_000
    between_answer_clicks_ms: int = 200

    # Настройки поведения
    fill_form_in_ui: bool = True
    auto_save_after_fill: bool = True
    manual_review_after_fill: bool = False

    # Попап-комментарии могут быть тяжёлыми по DOM
    load_popover_comments: bool = False
    load_lesson_popup_comments: bool = False


def build_default_config() -> AppConfig:
    browser_path = find_yandex_browser_path()
    if browser_path is None:
        browser_path = get_yandex_browser_candidates()[0]

    return AppConfig(
        yandex_browser_path=browser_path,
        user_data_dir=get_default_user_data_dir(),
        profile_directory="Default",
        xlsx_path=BASE_DIR / "student_skills_recommendations.xlsx",
    )


YANDEX_BROWSER_PATH = ""
USER_DATA_DIR = ""
PROFILE_DIRECTORY = "Default"
XLSX_PATH = BASE_DIR / "student_skills_recommendations.xlsx"

HEADLESS = False
MAX_CONCURRENT_JOURNALS = 1
REQUEST_TIMEOUT = 60_000
SAVE_ENABLE_TIMEOUT = 12_000

JOURNAL_EXTRA_STABILIZATION_MS = 1_000
BEFORE_STUDENT_CLICK_MS = 1_000
AFTER_STUDENT_POPUP_VISIBLE_MS = 1_000
AFTER_SIDE_PAGE_VISIBLE_MS = 1_000
BETWEEN_ANSWER_CLICKS_MS = 200

FILL_FORM_IN_UI = True
AUTO_SAVE_AFTER_FILL = True
MANUAL_REVIEW_AFTER_FILL = False

LOAD_POPOVER_COMMENTS = False
LOAD_LESSON_POPUP_COMMENTS = False


def apply_config(config: AppConfig) -> None:
    global YANDEX_BROWSER_PATH
    global USER_DATA_DIR
    global PROFILE_DIRECTORY
    global XLSX_PATH
    global HEADLESS
    global MAX_CONCURRENT_JOURNALS
    global REQUEST_TIMEOUT
    global SAVE_ENABLE_TIMEOUT
    global JOURNAL_EXTRA_STABILIZATION_MS
    global BEFORE_STUDENT_CLICK_MS
    global AFTER_STUDENT_POPUP_VISIBLE_MS
    global AFTER_SIDE_PAGE_VISIBLE_MS
    global BETWEEN_ANSWER_CLICKS_MS
    global FILL_FORM_IN_UI
    global AUTO_SAVE_AFTER_FILL
    global MANUAL_REVIEW_AFTER_FILL
    global LOAD_POPOVER_COMMENTS
    global LOAD_LESSON_POPUP_COMMENTS

    YANDEX_BROWSER_PATH = str(config.yandex_browser_path)
    USER_DATA_DIR = str(config.user_data_dir)
    PROFILE_DIRECTORY = config.profile_directory or "Default"
    XLSX_PATH = config.xlsx_path

    HEADLESS = config.headless
    MAX_CONCURRENT_JOURNALS = config.max_concurrent_journals
    REQUEST_TIMEOUT = config.request_timeout
    SAVE_ENABLE_TIMEOUT = config.save_enable_timeout

    JOURNAL_EXTRA_STABILIZATION_MS = config.journal_extra_stabilization_ms
    BEFORE_STUDENT_CLICK_MS = config.before_student_click_ms
    AFTER_STUDENT_POPUP_VISIBLE_MS = config.after_student_popup_visible_ms
    AFTER_SIDE_PAGE_VISIBLE_MS = config.after_side_page_visible_ms
    BETWEEN_ANSWER_CLICKS_MS = config.between_answer_clicks_ms

    FILL_FORM_IN_UI = config.fill_form_in_ui
    AUTO_SAVE_AFTER_FILL = config.auto_save_after_fill
    MANUAL_REVIEW_AFTER_FILL = config.manual_review_after_fill

    LOAD_POPOVER_COMMENTS = config.load_popover_comments
    LOAD_LESSON_POPUP_COMMENTS = config.load_lesson_popup_comments


APP_CONFIG = build_default_config()
apply_config(APP_CONFIG)

ASSESS_BUTTON_RE = re.compile(r"^\s*Оценить учебные умения\s*$", re.IGNORECASE)
REASSESS_BUTTON_RE = re.compile(r"^\s*Повторно оценить учебные умения\s*$", re.IGNORECASE)

ANSWER_LABELS = [
    "Почти никогда",
    "Редко",
    "Часто",
    "Почти всегда",
]

EMPTY_ANSWERS = [""] * 10

QUESTION_ORDER_HINT = [
    "Систематически выполняет домашние задания, готовится к урокам, имеет все нужное для занятий, рабочее место организовано",
    "Выполняет задания в срок, правильно распределяет время дома и на уроке",
    "Интересуется вопросами за рамками школьной программы, оперирует знаниями, полученными самостоятельно, приводит примеры не из школьной практики",
    "Интересуется более сложными задачами, проявляет любознательность",
    "Работает внимательно, легко переключается между разными задачами",
    "Оценивает свою работу и результаты объективно",
    "Регулирует свое поведение, соблюдает дисциплину",
    "Умеет слушать собеседника, легко поддерживает разговор, не перебивает",
    "Общается вежливо, соблюдает правила и нормы поведения",
    "Уверенно выступает перед аудиторией",
]

STATUS_SUCCESS = "успешно сохранен"
STATUS_OUU_INACTIVE = "кнопка ОУУ не активна"
STATUS_SAVE_INACTIVE = "кнопка сохранения не активна"
STATUS_OTHER_ERROR = "другая ошибка (вероятно Вы параллельно работали с журналом)"

ALL_STATUSES = [
    STATUS_SUCCESS,
    STATUS_OUU_INACTIVE,
    STATUS_SAVE_INACTIVE,
    STATUS_OTHER_ERROR,
]

MAIN_HEADERS = [
    "Параллель",
    "Класс",
    "Предмет",
    "Фамилия_Имя",
    "Отметки_и_комментарии",
    "Критерии",
    "Статус",
] + [f"{QUESTION_ORDER_HINT[i][:20]}" for i in range(10)]

SUMMARY_HEADERS = [
    "Ссылка_на_журнал",
    "Параллель",
    "Класс",
    "Предмет",
    "Общее_количество_обучающихся",
    "Успешно_сохранен",
    "Кнопка_ОУУ_не_активна",
    "Кнопка_сохранения_не_активна",
    "Другая_ошибка",
]


class ToolTip:
    def __init__(self, widget, text: str):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, _event=None):
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            justify="left",
            background="#fff8dc",
            relief="solid",
            borderwidth=1,
            padx=8,
            pady=5,
            wraplength=420,
        )
        label.pack()

    def hide(self, _event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


def show_startup_dialog(initial: AppConfig) -> Optional[AppConfig]:
    root = tk.Tk()
    root.title(APP_TITLE)
    root.geometry("980x760")
    root.minsize(900, 700)

    result: Dict[str, Optional[AppConfig]] = {"config": None}

    vars_map = {
        "yandex_browser_path": tk.StringVar(value=str(initial.yandex_browser_path)),
        "user_data_dir": tk.StringVar(value=str(initial.user_data_dir)),
        "profile_directory": tk.StringVar(value=initial.profile_directory),
        "xlsx_path": tk.StringVar(value=str(initial.xlsx_path)),
        "headless": tk.BooleanVar(value=initial.headless),
        "max_concurrent_journals": tk.StringVar(value=str(initial.max_concurrent_journals)),
        "request_timeout": tk.StringVar(value=str(initial.request_timeout)),
        "save_enable_timeout": tk.StringVar(value=str(initial.save_enable_timeout)),
        "journal_extra_stabilization_ms": tk.StringVar(value=str(initial.journal_extra_stabilization_ms)),
        "before_student_click_ms": tk.StringVar(value=str(initial.before_student_click_ms)),
        "after_student_popup_visible_ms": tk.StringVar(value=str(initial.after_student_popup_visible_ms)),
        "after_side_page_visible_ms": tk.StringVar(value=str(initial.after_side_page_visible_ms)),
        "between_answer_clicks_ms": tk.StringVar(value=str(initial.between_answer_clicks_ms)),
        "fill_form_in_ui": tk.BooleanVar(value=initial.fill_form_in_ui),
        "auto_save_after_fill": tk.BooleanVar(value=initial.auto_save_after_fill),
        "manual_review_after_fill": tk.BooleanVar(value=initial.manual_review_after_fill),
        "load_popover_comments": tk.BooleanVar(value=initial.load_popover_comments),
        "load_lesson_popup_comments": tk.BooleanVar(value=initial.load_lesson_popup_comments),
    }

    descriptions = {
        "yandex_browser_path": "Путь к browser.exe установленного Яндекс.Браузера. Сам браузер в exe не упаковывается и должен быть установлен у пользователя.",
        "user_data_dir": "Пользовательская директория браузера. По умолчанию используется профиль текущего пользователя: %LOCALAPPDATA%\\Yandex\\YandexBrowser\\User Data.",
        "profile_directory": "Имя профиля внутри User Data. Обычно используется Default.",
        "xlsx_path": "Путь к Excel-файлу результата. Если файла нет, он будет создан автоматически.",
        "headless": "Скрытый запуск браузера. Для пользовательского профиля обычно лучше оставлять выключенным.",
        "max_concurrent_journals": "Сколько журналов обрабатывать одновременно.",
        "request_timeout": "Общий таймаут ожиданий Playwright в миллисекундах.",
        "save_enable_timeout": "Сколько ждать активации кнопки сохранения после заполнения оценивания.",
        "journal_extra_stabilization_ms": "Дополнительная пауза после открытия журнала, чтобы страница стабилизировалась.",
        "before_student_click_ms": "Пауза перед кликом по ученику.",
        "after_student_popup_visible_ms": "Пауза после появления попапа ученика.",
        "after_side_page_visible_ms": "Пауза после появления боковой панели с вопросами.",
        "between_answer_clicks_ms": "Пауза между кликами по вариантам ответов.",
        "fill_form_in_ui": "Заполнять форму ОУУ прямо в интерфейсе журнала.",
        "auto_save_after_fill": "После заполнения автоматически нажимать кнопку сохранения.",
        "manual_review_after_fill": "После заполнения ждать ручной проверки перед продолжением.",
        "load_popover_comments": "Считывать всплывающие комментарии по ученикам и ячейкам. Может замедлить работу.",
        "load_lesson_popup_comments": "Считывать всплывающие комментарии из заголовков уроков. Может замедлить работу.",
    }

    def add_info_icon(parent, row: int, col: int, key: str):
        lbl = tk.Label(parent, text="ⓘ", fg="#005bbb", cursor="question_arrow")
        lbl.grid(row=row, column=col, sticky="w", padx=(6, 0))
        ToolTip(lbl, descriptions[key])

    def browse_file(var: tk.StringVar, filetypes):
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            var.set(filename)

    def browse_save_file(var: tk.StringVar):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=Path(var.get()).name if var.get() else "student_skills_recommendations.xlsx",
        )
        if filename:
            var.set(filename)

    def browse_directory(var: tk.StringVar):
        dirname = filedialog.askdirectory(initialdir=var.get() or str(BASE_DIR))
        if dirname:
            var.set(dirname)

    def add_entry_row(parent, row: int, title: str, key: str, browse_cmd=None):
        label = ttk.Label(parent, text=title)
        label.grid(row=row, column=0, sticky="w", padx=8, pady=4)
        entry = ttk.Entry(parent, textvariable=vars_map[key], width=82)
        entry.grid(row=row, column=1, sticky="ew", padx=8, pady=4)
        ToolTip(label, descriptions[key])
        ToolTip(entry, descriptions[key])
        if browse_cmd is not None:
            ttk.Button(parent, text="...", width=4, command=browse_cmd).grid(row=row, column=2, padx=(0, 4), pady=4)
        add_info_icon(parent, row, 3, key)

    def add_bool_row(parent, row: int, title: str, key: str):
        cb = ttk.Checkbutton(parent, text=title, variable=vars_map[key])
        cb.grid(row=row, column=0, columnspan=3, sticky="w", padx=8, pady=4)
        ToolTip(cb, descriptions[key])
        add_info_icon(parent, row, 3, key)

    def add_int_row(parent, row: int, title: str, key: str):
        label = ttk.Label(parent, text=title)
        label.grid(row=row, column=0, sticky="w", padx=8, pady=4)
        entry = ttk.Entry(parent, textvariable=vars_map[key], width=20)
        entry.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        ToolTip(label, descriptions[key])
        ToolTip(entry, descriptions[key])
        add_info_icon(parent, row, 3, key)

    container = ttk.Frame(root, padding=12)
    container.pack(fill="both", expand=True)
    container.columnconfigure(0, weight=1)

    top_hint = ttk.Label(
        container,
        text="Проверьте параметры запуска. Для подсказки наведите курсор на значок ⓘ.",
    )
    top_hint.grid(row=0, column=0, sticky="w", pady=(0, 10))

    frame_paths = ttk.LabelFrame(container, text="Пути")
    frame_paths.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
    frame_paths.columnconfigure(1, weight=1)

    add_entry_row(
        frame_paths,
        0,
        "Яндекс.Браузер (browser.exe)",
        "yandex_browser_path",
        browse_cmd=lambda: browse_file(
            vars_map["yandex_browser_path"],
            [("Executable files", "*.exe"), ("All files", "*.*")],
        ),
    )
    add_entry_row(
        frame_paths,
        1,
        "User Data",
        "user_data_dir",
        browse_cmd=lambda: browse_directory(vars_map["user_data_dir"]),
    )
    add_entry_row(frame_paths, 2, "Profile directory", "profile_directory")
    add_entry_row(
        frame_paths,
        3,
        "Excel-файл результата",
        "xlsx_path",
        browse_cmd=lambda: browse_save_file(vars_map["xlsx_path"]),
    )

    frame_timeouts = ttk.LabelFrame(container, text="Таймауты и паузы")
    frame_timeouts.grid(row=2, column=0, sticky="nsew", pady=(0, 10))
    frame_timeouts.columnconfigure(1, weight=1)

    add_int_row(frame_timeouts, 0, "Таймаут запросов (мс)", "request_timeout")
    add_int_row(frame_timeouts, 1, "Ожидание активации сохранения (мс)", "save_enable_timeout")
    add_int_row(frame_timeouts, 2, "Стабилизация журнала (мс)", "journal_extra_stabilization_ms")
    add_int_row(frame_timeouts, 3, "Перед кликом по ученику (мс)", "before_student_click_ms")
    add_int_row(frame_timeouts, 4, "После открытия попапа ученика (мс)", "after_student_popup_visible_ms")
    add_int_row(frame_timeouts, 5, "После открытия боковой панели (мс)", "after_side_page_visible_ms")
    add_int_row(frame_timeouts, 6, "Между ответами (мс)", "between_answer_clicks_ms")
    add_int_row(frame_timeouts, 7, "Параллельных журналов", "max_concurrent_journals")

    frame_behavior = ttk.LabelFrame(container, text="Поведение")
    frame_behavior.grid(row=3, column=0, sticky="nsew", pady=(0, 10))
    add_bool_row(frame_behavior, 0, "Headless режим", "headless")
    add_bool_row(frame_behavior, 1, "Заполнять форму в интерфейсе", "fill_form_in_ui")
    add_bool_row(frame_behavior, 2, "Автосохранение после заполнения", "auto_save_after_fill")
    add_bool_row(frame_behavior, 3, "Ручная проверка перед продолжением", "manual_review_after_fill")
    add_bool_row(frame_behavior, 4, "Загружать popover-комментарии", "load_popover_comments")
    add_bool_row(frame_behavior, 5, "Загружать комментарии из уроков", "load_lesson_popup_comments")

    buttons = ttk.Frame(container)
    buttons.grid(row=4, column=0, sticky="e", pady=(6, 0))

    def parse_non_negative_int(title: str, value: str) -> int:
        try:
            parsed = int(value)
        except Exception:
            raise ValueError(f"Поле '{title}' должно быть целым числом.")
        if parsed < 0:
            raise ValueError(f"Поле '{title}' не может быть отрицательным.")
        return parsed

    def on_start():
        try:
            browser_raw = vars_map["yandex_browser_path"].get().strip()
            browser_path = clean_path_from_ui(browser_raw) if browser_raw else (find_yandex_browser_path() or Path())
            if not browser_path.exists() or not browser_path.is_file():
                raise ValueError("Не найден browser.exe Яндекс.Браузера. Укажите корректный путь.")

            user_data_dir = clean_path_from_ui(vars_map["user_data_dir"].get())
            user_data_dir.mkdir(parents=True, exist_ok=True)

            xlsx_path = clean_path_from_ui(vars_map["xlsx_path"].get())
            if not xlsx_path.is_absolute():
                xlsx_path = (BASE_DIR / xlsx_path).resolve()
            xlsx_path.parent.mkdir(parents=True, exist_ok=True)

            config = AppConfig(
                yandex_browser_path=browser_path.resolve(),
                user_data_dir=user_data_dir.resolve(),
                profile_directory=vars_map["profile_directory"].get().strip() or "Default",
                xlsx_path=xlsx_path,
                headless=vars_map["headless"].get(),
                max_concurrent_journals=parse_non_negative_int("Параллельных журналов", vars_map["max_concurrent_journals"].get()),
                request_timeout=parse_non_negative_int("Таймаут запросов (мс)", vars_map["request_timeout"].get()),
                save_enable_timeout=parse_non_negative_int("Ожидание активации сохранения (мс)", vars_map["save_enable_timeout"].get()),
                journal_extra_stabilization_ms=parse_non_negative_int("Стабилизация журнала (мс)", vars_map["journal_extra_stabilization_ms"].get()),
                before_student_click_ms=parse_non_negative_int("Перед кликом по ученику (мс)", vars_map["before_student_click_ms"].get()),
                after_student_popup_visible_ms=parse_non_negative_int("После открытия попапа ученика (мс)", vars_map["after_student_popup_visible_ms"].get()),
                after_side_page_visible_ms=parse_non_negative_int("После открытия боковой панели (мс)", vars_map["after_side_page_visible_ms"].get()),
                between_answer_clicks_ms=parse_non_negative_int("Между ответами (мс)", vars_map["between_answer_clicks_ms"].get()),
                fill_form_in_ui=vars_map["fill_form_in_ui"].get(),
                auto_save_after_fill=vars_map["auto_save_after_fill"].get(),
                manual_review_after_fill=vars_map["manual_review_after_fill"].get(),
                load_popover_comments=vars_map["load_popover_comments"].get(),
                load_lesson_popup_comments=vars_map["load_lesson_popup_comments"].get(),
            )
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc), parent=root)
            return

        result["config"] = config
        root.destroy()

    def on_cancel():
        result["config"] = None
        root.destroy()

    ttk.Button(buttons, text="Запустить", command=on_start).pack(side="left", padx=(0, 8))
    ttk.Button(buttons, text="Отмена", command=on_cancel).pack(side="left")

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return result["config"]


class SaveButtonInactiveError(RuntimeError):
    pass


@dataclass
class JournalEntry:
    parallel: str
    class_name: str
    subject: str
    url: str


@dataclass
class LocalStudentData:
    name: str
    marks: List[List[str]] = field(default_factory=list)  # [mark, form, comment]
    comments: List[str] = field(default_factory=list)
    lesson_comments: List[str] = field(default_factory=list)
    n_list: List[str] = field(default_factory=list)
    numeric_marks: List[float] = field(default_factory=list)
    mark_count: int = 0
    source_url: str = ""
    parallel: str = ""
    class_name: str = ""
    subject: str = ""


@dataclass
class StudentAggregate:
    name: str
    parallels: set = field(default_factory=set)
    classes: set = field(default_factory=set)
    subjects: set = field(default_factory=set)
    source_urls: List[str] = field(default_factory=list)

    marks: List[List[str]] = field(default_factory=list)
    comments: List[str] = field(default_factory=list)
    lesson_comments: List[str] = field(default_factory=list)
    n_list: List[str] = field(default_factory=list)
    numeric_marks: List[float] = field(default_factory=list)

    mark_density_ratios: List[float] = field(default_factory=list)

    first_parallel: str = ""
    first_class: str = ""
    first_subject: str = ""
    first_url: str = ""


def agg_from_local(s: LocalStudentData) -> StudentAggregate:
    agg = StudentAggregate(
        name=s.name,
        first_parallel=s.parallel,
        first_class=s.class_name,
        first_subject=s.subject,
        first_url=s.source_url,
    )
    agg.parallels.add(s.parallel)
    agg.classes.add(s.class_name)
    agg.subjects.add(s.subject)
    if s.source_url:
        agg.source_urls.append(s.source_url)

    agg.marks = list(s.marks)
    agg.comments = list(s.comments)
    agg.lesson_comments = list(s.lesson_comments)
    agg.n_list = list(s.n_list)
    agg.numeric_marks = list(s.numeric_marks)
    agg.mark_density_ratios = [1.0]
    return agg


def norm(s: Optional[str]) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def uniq_keep_order(items: List[str]) -> List[str]:
    seen = set()
    out = []
    for item in items:
        item = norm(item)
        if not item or item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def safe_mean(values: List[float], default: float = 0.0) -> float:
    vals = [v for v in values if v is not None]
    if not vals:
        return default
    return sum(vals) / len(vals)


def clamp(v: int, lo: int = 0, hi: int = 3) -> int:
    return max(lo, min(hi, v))


def json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, sort_keys=False)


def to_float_mark(mark: str) -> Optional[float]:
    if mark is None:
        return None
    s = str(mark).strip().replace(",", ".")
    if re.fullmatch(r"\d+(?:\.\d+)?", s):
        try:
            return float(s)
        except Exception:
            return None
    return None


def parse_mark_cell_text(text: str) -> Dict[str, Any]:
    text = norm(text)
    if not text:
        return {
            "raw": "",
            "mark": None,
            "absent": False,
            "flags": [],
        }

    tokens = [t.strip() for t in re.split(r"\s+", text) if t.strip()]
    lower_tokens = [t.lower() for t in tokens]

    absent = "н" in lower_tokens
    flags = [t for t in lower_tokens if t in {"к", "б"}]

    mark_token = None
    for token in tokens:
        t = token.lower()
        if t in {"н", "к", "б"}:
            continue
        mark_token = token
        break

    return {
        "raw": text,
        "mark": mark_token,
        "absent": absent,
        "flags": flags,
    }


def count_keyword_hits(text: str, patterns: List[str]) -> int:
    text = text.lower()
    total = 0
    for p in patterns:
        total += len(re.findall(p, text))
    return total


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def ensure_main_sheet(ws) -> None:
    if ws.max_row < 1:
        ws.append(MAIN_HEADERS)
        return

    header_values = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    if "Статус" not in header_values:
        ws.insert_cols(7)

    for idx, header in enumerate(MAIN_HEADERS, start=1):
        ws.cell(1, idx).value = header


def ensure_summary_sheet(ws) -> None:
    if ws.max_row < 1:
        ws.append(SUMMARY_HEADERS)
        return

    for idx, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(1, idx).value = header


def ensure_workbook(path: Path) -> Tuple[Workbook, Any, Any]:
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    if "Рекомендации" in wb.sheetnames:
        ws_main = wb["Рекомендации"]
    else:
        ws_main = wb.active
        ws_main.title = "Рекомендации"

    ensure_main_sheet(ws_main)

    if "Сводка" in wb.sheetnames:
        ws_summary = wb["Сводка"]
    else:
        ws_summary = wb.create_sheet("Сводка")

    ensure_summary_sheet(ws_summary)
    wb.save(path)
    return wb, ws_main, ws_summary


def get_header_index_map(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = norm(ws.cell(1, col).value)
        if key:
            out[key] = col
    return out


def load_processed_names(ws) -> Set[str]:
    processed = set()
    headers = get_header_index_map(ws)

    name_col = headers.get("Фамилия_Имя")
    status_col = headers.get("Статус")
    if not name_col or not status_col:
        return processed

    for row in ws.iter_rows(min_row=2):
        try:
            name = norm(row[name_col - 1].value)
            status = norm(row[status_col - 1].value).lower()
        except Exception:
            continue

        if name and status == STATUS_SUCCESS:
            processed.add(name)

    return processed


def load_summary_completion_map(ws_summary) -> Dict[str, Dict[str, int]]:
    result: Dict[str, Dict[str, int]] = {}
    for row in ws_summary.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        url = norm(row[0] if len(row) > 0 else "")
        if not url:
            continue
        result[url] = {
            "total": safe_int(row[4] if len(row) > 4 else 0),
            "success": safe_int(row[5] if len(row) > 5 else 0),
            "ouu_inactive": safe_int(row[6] if len(row) > 6 else 0),
            "save_inactive": safe_int(row[7] if len(row) > 7 else 0),
            "other_error": safe_int(row[8] if len(row) > 8 else 0),
        }
    return result


def append_student_row(
    ws,
    agg: StudentAggregate,
    criteria: Dict[str, Any],
    status: str,
    answers: List[str],
) -> None:
    payload = {
        "marks": agg.marks,
        "comments": uniq_keep_order(agg.comments),
        "lesson_comments": uniq_keep_order(agg.lesson_comments),
        "n_list": agg.n_list,
        "all_subjects": sorted(agg.subjects),
        "all_classes": sorted(agg.classes),
        "all_parallels": sorted(agg.parallels),
        "all_source_urls": agg.source_urls,
    }

    row = [
        ", ".join(sorted(agg.parallels)),
        ", ".join(sorted(agg.classes)),
        ", ".join(sorted(agg.subjects)),
        agg.name,
        json_dumps(payload),
        json_dumps(criteria),
        status,
    ] + answers

    ws.append(row)


def upsert_summary_row(
    ws_summary,
    entry: JournalEntry,
    total_students: int,
    status_counts: Dict[str, int],
) -> None:
    target_row = None
    for row_idx in range(2, ws_summary.max_row + 1):
        url = norm(ws_summary.cell(row_idx, 1).value)
        if url == entry.url:
            target_row = row_idx
            break

    if target_row is None:
        target_row = ws_summary.max_row + 1

    ws_summary.cell(target_row, 1).value = entry.url
    ws_summary.cell(target_row, 2).value = entry.parallel
    ws_summary.cell(target_row, 3).value = entry.class_name
    ws_summary.cell(target_row, 4).value = entry.subject
    ws_summary.cell(target_row, 5).value = total_students
    ws_summary.cell(target_row, 6).value = status_counts.get(STATUS_SUCCESS, 0)
    ws_summary.cell(target_row, 7).value = status_counts.get(STATUS_OUU_INACTIVE, 0)
    ws_summary.cell(target_row, 8).value = status_counts.get(STATUS_SAVE_INACTIVE, 0)
    ws_summary.cell(target_row, 9).value = status_counts.get(STATUS_OTHER_ERROR, 0)


async def init_browser(headless: bool = False) -> Tuple[Playwright, BrowserContext, Page]:
    playwright = await async_playwright().start()

    executable_path = Path(YANDEX_BROWSER_PATH)
    if not executable_path.exists():
        detected = find_yandex_browser_path()
        if detected is None or not detected.exists():
            raise FileNotFoundError("Не найден установленный Яндекс.Браузер. Укажите корректный путь к browser.exe.")
        executable_path = detected

    user_data_dir = Path(USER_DATA_DIR).expanduser()
    user_data_dir.mkdir(parents=True, exist_ok=True)

    launch_args = []
    if PROFILE_DIRECTORY:
        launch_args.append(f"--profile-directory={PROFILE_DIRECTORY}")

    context = await playwright.chromium.launch_persistent_context(
        user_data_dir=str(user_data_dir),
        executable_path=str(executable_path),
        headless=headless,
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        args=launch_args,
    )

    page = context.pages[0] if context.pages else await context.new_page()
    return playwright, context, page


async def safe_goto(page: Page, url: str) -> None:
    await page.goto(url, wait_until="domcontentloaded", timeout=REQUEST_TIMEOUT)


async def wait_visible(locator) -> bool:
    try:
        await locator.wait_for(state="visible", timeout=REQUEST_TIMEOUT)
        return True
    except Exception:
        return False


async def wait_user_input(message: str) -> None:
    await asyncio.to_thread(input, message)


async def click_locator(locator, timeout: int = 3000, force: bool = False) -> bool:
    try:
        if await locator.count() == 0:
            return False
        try:
            await locator.first.scroll_into_view_if_needed()
        except Exception:
            pass
        await locator.first.click(timeout=timeout, force=force)
        return True
    except Exception:
        return False


async def locator_enabled_now(locator) -> bool:
    try:
        if await locator.count() == 0:
            return False
        el = locator.first
        if not await el.is_visible():
            return False
        if not await el.is_enabled():
            return False

        aria_disabled = await el.get_attribute("aria-disabled")
        disabled_attr = await el.get_attribute("disabled")
        if aria_disabled in {"true", "1"}:
            return False
        if disabled_attr is not None:
            return False
        return True
    except Exception:
        return False


async def wait_locator_enabled(locator, timeout: int = SAVE_ENABLE_TIMEOUT, poll_ms: int = 200) -> bool:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + timeout / 1000

    while loop.time() < deadline:
        if await locator_enabled_now(locator):
            return True
        await asyncio.sleep(poll_ms / 1000)

    return False


async def expand_all_parallel_sections(page: Page) -> None:
    await page.wait_for_selector("section[data-test-component^='journalListSection-']", timeout=REQUEST_TIMEOUT)

    titles = [f"{i} параллель" for i in range(1, 12)] + ["Смешанные группы"]

    for title in titles:
        section = page.locator("section[data-test-component^='journalListSection-']").filter(
            has=page.locator("h6", has_text=title)
        )
        if await section.count() == 0:
            continue

        sec = section.first
        cards = sec.locator("div[data-test-component='journalCardRoute']")
        if await cards.count() > 0:
            continue

        candidates = [
            sec.locator("h6", has_text=title).first,
            sec.locator(":scope > div > div").first,
            sec.locator("svg").first,
        ]

        for loc in candidates:
            try:
                if await loc.count() == 0:
                    continue
                await loc.scroll_into_view_if_needed()
                await loc.click(timeout=2000)
                await page.wait_for_timeout(500)
                break
            except Exception:
                continue

        await page.wait_for_timeout(300)


async def scrape_journal_entries(page: Page) -> List[JournalEntry]:
    await safe_goto(page, JOURNALS_URL)
    await page.wait_for_selector("section[data-test-component^='journalListSection-']", timeout=REQUEST_TIMEOUT)
    await expand_all_parallel_sections(page)

    sections = page.locator("section[data-test-component^='journalListSection-']")
    entries: List[JournalEntry] = []
    seen = set()

    for i in range(await sections.count()):
        section = sections.nth(i)

        try:
            parallel = norm(await section.locator("h6").first.inner_text())
        except Exception:
            parallel = ""

        cards = section.locator("div[data-test-component='journalCardRoute']")
        for j in range(await cards.count()):
            card = cards.nth(j)

            href = await card.locator("a").first.get_attribute("href")
            if not href:
                continue

            full_url = urljoin(BASE_URL, href)
            if full_url in seen:
                continue

            texts = [norm(t) for t in await card.locator("span").all_inner_texts()]
            texts = [t for t in texts if t]

            if len(texts) >= 2:
                class_name = texts[0]
                subject = texts[-1]
            elif len(texts) == 1:
                class_name = texts[0]
                subject = ""
            else:
                class_name = ""
                subject = ""

            entries.append(
                JournalEntry(
                    parallel=parallel,
                    class_name=class_name,
                    subject=subject,
                    url=full_url,
                )
            )
            seen.add(full_url)

    return entries


async def ensure_all_periods_selected(page: Page) -> None:
    btn = page.locator("[data-test-component='JournalControlPanelFilters']")
    if await btn.count() == 0:
        return

    try:
        await btn.first.click(timeout=3000)
    except Exception:
        return

    popup = page.locator("[data-test-component='journalControlPanelFiltersPopup']")
    if not await wait_visible(popup):
        return

    checkboxes = popup.locator("li[data-test-component^='JournalControlPanelFiltersPopupTopic-'] input[type='checkbox']")
    for i in range(await checkboxes.count()):
        cb = checkboxes.nth(i)
        try:
            checked = await cb.is_checked()
            if not checked:
                await cb.check(force=True)
        except Exception:
            try:
                await cb.click(force=True)
            except Exception:
                pass

    apply_btn = popup.locator("[data-test-component='JournalControlPanelFiltersPopupSave']")
    if await apply_btn.count() > 0:
        try:
            await apply_btn.first.click()
        except Exception:
            try:
                await apply_btn.first.click(force=True)
            except Exception:
                pass

    await page.wait_for_timeout(2000)


EXTRACT_TABLE_JS = r"""
() => {
    const table = document.querySelector("table");
    if (!table) return null;

    const headerMap = {};
    table.querySelectorAll("thead [data-test-component^='controlFormContentHeadCell-']").forEach(el => {
        const raw = el.getAttribute("data-test-component") || "";
        const key = raw.replace("controlFormContentHeadCell-", "");
        headerMap[key] = (el.innerText || "").replace(/\s+/g, " ").trim();
    });

    const lessonHeaders = [];
    table.querySelectorAll("thead [data-test-component^='scheduleLessonCell-']").forEach(el => {
        const raw = el.getAttribute("data-test-component") || "";
        const txt = (el.innerText || "").replace(/\s+/g, " ").trim();
        lessonHeaders.push({ testid: raw, text: txt });
    });

    const students = [];
    const rows = Array.from(table.querySelectorAll("tbody tr"));

    for (const tr of rows) {
        const nameEl = tr.querySelector("td span[title]");
        if (!nameEl) continue;

        const name = (nameEl.getAttribute("title") || nameEl.innerText || "").replace(/\s+/g, " ").trim();
        if (!name) continue;

        const firstCell = tr.querySelector("td");
        const studentCommentNode = firstCell ? firstCell.querySelector("[data-test-component^='studentCellInfoComments-']") : null;
        const studentCommentTestId = studentCommentNode ? studentCommentNode.getAttribute("data-test-component") : null;
        const studentCommentHasContent = !!(studentCommentNode && studentCommentNode.children && studentCommentNode.children.length > 0);

        const cells = [];
        tr.querySelectorAll("td [data-test-component^='markCell-']").forEach(cell => {
            const testid = cell.getAttribute("data-test-component") || "";
            const text = (cell.innerText || "").replace(/\s+/g, " ").trim();
            const hasSvg = !!cell.querySelector("svg");
            cells.push({
                testid,
                text,
                hasSvg
            });
        });

        students.push({
            name,
            studentCommentTestId,
            studentCommentHasContent,
            cells
        });
    }

    return {
        headerMap,
        lessonHeaders,
        students
    };
}
"""


async def extract_table_basic(page: Page) -> Dict[str, Any]:
    await page.wait_for_selector("table", timeout=REQUEST_TIMEOUT)
    data = await page.evaluate(EXTRACT_TABLE_JS)
    if not data:
        raise RuntimeError("Не удалось получить данные таблицы журнала")
    return data


async def get_visible_popup_texts(page: Page) -> List[str]:
    return await page.evaluate(
        r"""
() => {
    const selectors = [
        "div.MuiPopper-root",
        "[role='tooltip']",
        "[data-test-component*='Popover']",
        "[data-test-component*='popover']"
    ];

    const out = [];
    const seen = new Set();

    for (const el of document.querySelectorAll(selectors.join(","))) {
        const style = window.getComputedStyle(el);
        const rect = el.getBoundingClientRect();
        if (style.display === "none" || style.visibility === "hidden" || rect.width === 0 || rect.height === 0) {
            continue;
        }
        const txt = (el.innerText || "").replace(/\s+/g, " ").trim();
        if (!txt) continue;
        if (txt.length > 1500) continue;
        if (!seen.has(txt)) {
            seen.add(txt);
            out.push(txt);
        }
    }
    return out;
}
"""
    )


def clean_popup_texts(texts: List[str]) -> List[str]:
    cleaned = []
    for t in texts:
        t = norm(t)
        if not t:
            continue
        if t in {"Фильтры", "Информация об учащемся"}:
            continue
        cleaned.append(t)
    return uniq_keep_order(cleaned)


async def click_and_capture_popup_text(page: Page, locator) -> str:
    before = set(await get_visible_popup_texts(page))
    try:
        await locator.scroll_into_view_if_needed()
        await locator.click(force=True, timeout=2000)
        await page.wait_for_timeout(400)
    except Exception:
        return ""

    after = set(await get_visible_popup_texts(page))
    diff = [t for t in after if t not in before]
    diff = clean_popup_texts(diff)

    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(150)
    except Exception:
        pass

    return " | ".join(diff)


async def collect_optional_comments(page: Page, raw_table: Dict[str, Any]) -> Tuple[Dict[str, str], List[str]]:
    cell_comment_map: Dict[str, str] = {}
    lesson_comments: List[str] = []

    if LOAD_LESSON_POPUP_COMMENTS:
        for item in raw_table.get("lessonHeaders", []):
            testid = item.get("testid", "")
            if not testid:
                continue
            loc = page.locator(f"[data-test-component='{testid}']").first
            if await loc.count() == 0:
                continue
            txt = await click_and_capture_popup_text(page, loc)
            if txt:
                lesson_comments.append(txt)

    if LOAD_POPOVER_COMMENTS:
        for student in raw_table.get("students", []):
            if student.get("studentCommentHasContent") and student.get("studentCommentTestId"):
                tid = student["studentCommentTestId"]
                loc = page.locator(f"[data-test-component='{tid}']").first
                if await loc.count() > 0:
                    txt = await click_and_capture_popup_text(page, loc)
                    if txt:
                        cell_comment_map[tid] = txt

            for cell in student.get("cells", []):
                if not cell.get("hasSvg"):
                    continue
                tid = cell.get("testid", "")
                if not tid:
                    continue
                loc = page.locator(f"[data-test-component='{tid}']").first
                if await loc.count() == 0:
                    continue
                txt = await click_and_capture_popup_text(page, loc)
                if txt:
                    cell_comment_map[tid] = txt

    return cell_comment_map, uniq_keep_order(lesson_comments)


def build_local_students_from_table(
    entry: JournalEntry,
    raw_table: Dict[str, Any],
    popup_comment_map: Dict[str, str],
    journal_lesson_comments: List[str],
) -> List[LocalStudentData]:
    header_map: Dict[str, str] = raw_table.get("headerMap", {})
    students_raw: List[Dict[str, Any]] = raw_table.get("students", [])

    local_students: List[LocalStudentData] = []

    for student in students_raw:
        name = norm(student.get("name"))
        if not name:
            continue

        local = LocalStudentData(
            name=name,
            parallel=entry.parallel,
            class_name=entry.class_name,
            subject=entry.subject,
            source_url=entry.url,
        )

        if student.get("studentCommentHasContent"):
            tid = student.get("studentCommentTestId")
            if tid and tid in popup_comment_map:
                local.comments.append(popup_comment_map[tid])

        local.lesson_comments.extend(journal_lesson_comments)

        for cell in student.get("cells", []):
            testid = cell.get("testid", "")
            text = norm(cell.get("text", ""))

            if not testid:
                continue

            m = re.match(r"^markCell-(\d+)_([^_]+)_(.+)$", testid)
            if not m:
                continue

            col_id = m.group(2)
            suffix = m.group(3)

            if suffix == "finalResult":
                continue

            header_key = f"{col_id}_{suffix}"
            control_form = norm(header_map.get(header_key, ""))

            parsed = parse_mark_cell_text(text)
            comment = popup_comment_map.get(testid, "")

            if parsed["absent"]:
                local.n_list.append(control_form or header_key)
                if comment:
                    local.comments.append(comment)

            mark = parsed["mark"]
            if mark:
                local.marks.append([str(mark), control_form, comment])
                local.mark_count += 1

                num = to_float_mark(str(mark))
                if num is not None:
                    local.numeric_marks.append(num)
            else:
                if comment:
                    local.comments.append(comment)

        local.comments = uniq_keep_order(local.comments)
        local.lesson_comments = uniq_keep_order(local.lesson_comments)

        local_students.append(local)

    return local_students


def recommend_answers_for_student(agg: StudentAggregate) -> Tuple[List[str], Dict[str, Any]]:
    numeric = agg.numeric_marks[:]
    avg_mark = safe_mean(numeric, default=3.7)

    mark_count = len(agg.marks)
    n_count = len(agg.n_list)
    density_ratio = safe_mean(agg.mark_density_ratios, default=1.0)

    all_comments = uniq_keep_order(agg.comments + agg.lesson_comments)
    comments_text = " ".join(all_comments).lower()

    low_marks = sum(1 for x in numeric if x <= 2.5)
    high_marks = sum(1 for x in numeric if x >= 4.0)

    oral_numeric = []
    for mark, form, _comment in agg.marks:
        if norm(form).lower() == "устн":
            v = to_float_mark(mark)
            if v is not None:
                oral_numeric.append(v)

    oral_count = len(oral_numeric)
    oral_avg = safe_mean(oral_numeric, default=0.0)

    n_ratio = n_count / max(mark_count + n_count, 1)

    late_hits = count_keyword_hits(comments_text, [r"опозд"])
    phone_hits = count_keyword_hits(comments_text, [r"телефон", r"смартфон", r"в телеф"])
    play_hits = count_keyword_hits(comments_text, [r"игра", r"балует", r"балуется", r"болтает", r"разговар", r"мешает", r"отвлека", r"шумит"])
    rude_hits = count_keyword_hits(comments_text, [r"грубит", r"хамит", r"перебива"])
    no_work_hits = count_keyword_hits(comments_text, [r"не работ", r"не выполня", r"не готов"])
    positive_hits = count_keyword_hits(comments_text, [r"молодец", r"отлично", r"актив", r"стара", r"самостоят", r"инициатив"])

    behavior_hits = late_hits + phone_hits + play_hits + rude_hits + no_work_hits

    avg_shift = 1 if avg_mark >= 4.5 else (0 if avg_mark >= 3.4 else -1)
    density_shift = 1 if density_ratio >= 1.15 else (-1 if density_ratio < 0.75 else 0)

    if n_ratio >= 0.35 or n_count >= 10:
        absence_shift = -2
    elif n_ratio >= 0.15 or n_count >= 4:
        absence_shift = -1
    else:
        absence_shift = 0

    if behavior_hits >= 4:
        discipline_shift = -2
    elif behavior_hits >= 2:
        discipline_shift = -1
    elif behavior_hits == 0 and n_ratio < 0.12:
        discipline_shift = 1
    else:
        discipline_shift = 0

    if oral_count >= 2 and oral_avg >= 4.4:
        oral_shift = 1
    elif oral_count >= 2 and oral_avg < 3.5:
        oral_shift = -1
    else:
        oral_shift = 0

    consistency_shift = 0
    if len(numeric) >= 4:
        if (max(numeric) - min(numeric) <= 2.0) and avg_mark >= 4.0:
            consistency_shift = 1
        elif low_marks >= 4 and avg_mark < 3.3:
            consistency_shift = -1

    q1 = clamp(2 + avg_shift + density_shift + absence_shift + (-1 if late_hits >= 2 else 0))
    q2 = clamp(2 + density_shift + (1 if avg_mark >= 4.3 else 0) + (-1 if late_hits >= 1 else 0) + (-1 if n_count >= 4 else 0))
    q3 = clamp(2 + (1 if avg_mark >= 4.6 and density_ratio >= 1.0 else 0) + (1 if positive_hits >= 2 else 0) + (-1 if avg_mark < 3.2 else 0))
    q4 = clamp(2 + (1 if avg_mark >= 4.4 else 0) + (1 if density_ratio >= 1.15 else 0) + (-1 if avg_mark < 3.0 else 0))
    q5 = clamp(2 + avg_shift + (-1 if n_count >= 4 else 0) + (-1 if behavior_hits >= 2 else 0) + (1 if high_marks > max(low_marks * 2, 0) and avg_mark >= 4.0 else 0))
    q6 = clamp(2 + consistency_shift + (1 if avg_mark >= 4.3 and low_marks == 0 else 0) + (-1 if low_marks >= 4 and avg_mark < 3.3 else 0))
    q7 = clamp(2 + discipline_shift + (-1 if n_count >= 6 else 0))
    q8 = clamp(2 + oral_shift + (-1 if behavior_hits >= 3 else 0) + (1 if positive_hits >= 1 else 0))
    q9 = clamp(2 + (1 if behavior_hits == 0 else 0) + (-1 if phone_hits + play_hits >= 2 else 0) + (-1 if rude_hits >= 1 else 0))
    q10 = clamp(2 + oral_shift + (1 if oral_count >= 3 and oral_avg >= 4.6 else 0) + (-1 if oral_count >= 2 and oral_avg < 3.2 else 0))

    idx_answers = [q1, q2, q3, q4, q5, q6, q7, q8, q9, q10]

    if avg_mark >= 4.25 and low_marks == 0:
        floors = [3, 3, 3, 3, 3, 3, 2, 3, 2, 3]
        if behavior_hits < 2 and n_ratio < 0.20:
            idx_answers = [max(v, floors[i]) for i, v in enumerate(idx_answers)]
        else:
            idx_answers = [max(v, 2) for v in idx_answers]

    answers = [ANSWER_LABELS[i] for i in idx_answers]

    criteria = {
        "avg_mark": round(avg_mark, 3),
        "mark_count": mark_count,
        "n_count": n_count,
        "n_ratio": round(n_ratio, 3),
        "density_ratio": round(density_ratio, 3),
        "oral_count": oral_count,
        "oral_avg": round(oral_avg, 3) if oral_count else None,
        "low_marks": low_marks,
        "high_marks": high_marks,
        "late_hits": late_hits,
        "phone_hits": phone_hits,
        "play_hits": play_hits,
        "rude_hits": rude_hits,
        "no_work_hits": no_work_hits,
        "positive_hits": positive_hits,
        "behavior_hits": behavior_hits,
        "answers": {f"Q{i+1}": answers[i] for i in range(10)},
        "question_order_hint": QUESTION_ORDER_HINT,
        "note": "Рекомендация сформирована автоматически по объективным данным журнала. Итоговое решение и сохранение — за учителем.",
    }

    return answers, criteria


ANNOTATE_QUESTIONS_JS = r"""
() => {
    const side = document.querySelector("[data-test-component='undefined-sidePage']");
    if (!side) return [];

    const body = side.querySelector("[data-test-component='undefined-bodySidePage']");
    if (!body) return [];

    let idx = 0;
    const result = [];

    for (const section of Array.from(body.children)) {
        if (section.getAttribute("data-test-component")) {
            continue;
        }

        const sectionTitleEl = section.querySelector(":scope > span");
        if (!sectionTitleEl) continue;
        const sectionTitle = (sectionTitleEl.innerText || "").replace(/\s+/g, " ").trim();

        const children = Array.from(section.children).slice(1);
        for (const q of children) {
            const qTextEl = q.querySelector(":scope > span");
            if (!qTextEl) continue;

            const qText = (qTextEl.innerText || "").replace(/\s+/g, " ").trim();
            q.setAttribute("data-auto-question-index", String(idx));

            const options = Array.from(q.querySelectorAll("label")).map(l => (l.innerText || "").replace(/\s+/g, " ").trim());

            result.push({
                index: idx,
                section: sectionTitle,
                question: qText,
                options
            });
            idx += 1;
        }
    }

    return result;
}
"""


async def close_student_info_popup(page: Page, popup_locator=None) -> None:
    if popup_locator is None:
        popup_locator = page.locator("[role='tooltip']").filter(has=page.locator("text=Информация об учащемся")).last

    try:
        close_btn = popup_locator.locator(
            "[data-test-component='undefined-iconButtonInPopover-iconButtonComponent']"
        ).first
        if await click_locator(close_btn, timeout=3000, force=True):
            await page.wait_for_timeout(250)
            return
    except Exception:
        pass

    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(250)
    except Exception:
        pass


async def open_assessment_sidepage(page: Page, student_name: str) -> str:
    await page.wait_for_timeout(BEFORE_STUDENT_CLICK_MS)

    name_loc = page.locator("td span[title]").filter(has_text=student_name).first
    await name_loc.scroll_into_view_if_needed()
    await name_loc.click(timeout=5000)

    popup = page.locator("[role='tooltip']").filter(has=page.locator("text=Информация об учащемся"))
    await popup.last.wait_for(state="visible", timeout=REQUEST_TIMEOUT)

    await page.wait_for_timeout(AFTER_STUDENT_POPUP_VISIBLE_MS)

    current_popup = popup.last
    buttons = current_popup.locator("button")

    assess_btn = buttons.filter(has_text=ASSESS_BUTTON_RE)
    reassess_btn = buttons.filter(has_text=REASSESS_BUTTON_RE)

    if await assess_btn.count() > 0:
        if not await locator_enabled_now(assess_btn.first):
            await close_student_info_popup(page, current_popup)
            return "ouu_disabled"

        await assess_btn.first.click(timeout=5000)

        side = page.locator("[data-test-component='undefined-sidePage']")
        await side.locator("[data-test-component='undefined-infoBoxContainer']").wait_for(
            state="visible",
            timeout=REQUEST_TIMEOUT,
        )
        await page.wait_for_timeout(AFTER_SIDE_PAGE_VISIBLE_MS)
        return "new"

    if await reassess_btn.count() > 0:
        await close_student_info_popup(page, current_popup)
        return "already_assessed"

    await close_student_info_popup(page, current_popup)
    raise RuntimeError(
        f"У '{student_name}' не найдены ожидаемые кнопки "
        f"'Оценить учебные умения' / 'Повторно оценить учебные умения'"
    )


async def fill_assessment_answers(page: Page, answers: List[str]) -> List[Dict[str, Any]]:
    side = page.locator("[data-test-component='undefined-sidePage']")
    await side.wait_for(state="visible", timeout=REQUEST_TIMEOUT)

    question_meta = await page.evaluate(ANNOTATE_QUESTIONS_JS)

    total = min(len(question_meta), len(answers))
    for i in range(total):
        answer = answers[i]
        container = side.locator(f"[data-auto-question-index='{i}']").first
        label = container.locator("label", has_text=answer).first
        await label.scroll_into_view_if_needed()
        await label.click(timeout=2000)
        await page.wait_for_timeout(BETWEEN_ANSWER_CLICKS_MS)

    await page.wait_for_timeout(BETWEEN_ANSWER_CLICKS_MS)
    return question_meta


async def fill_recommendation_in_ui_on_current_journal(
    page: Page,
    student_name: str,
    answers: List[str],
) -> Tuple[str, List[Dict[str, Any]]]:
    status = await open_assessment_sidepage(page, student_name)

    if status == "already_assessed":
        return "already_assessed", []

    if status == "ouu_disabled":
        return "ouu_disabled", []

    q_meta = await fill_assessment_answers(page, answers)
    return "filled", q_meta


async def save_assessment_sidepage(page: Page) -> None:
    side = page.locator("[data-test-component='undefined-sidePage']")
    if not await wait_visible(side):
        raise RuntimeError("Боковая панель с оцениванием не найдена")

    await page.wait_for_timeout(BETWEEN_ANSWER_CLICKS_MS)

    primary_btn = side.locator("[data-test-component='undefined-footerSidePage-actionBarPrimaryButton']")
    enabled = await wait_locator_enabled(primary_btn, timeout=SAVE_ENABLE_TIMEOUT)
    if not enabled:
        raise SaveButtonInactiveError("Кнопка сохранения не стала активной после заполнения оценивания")

    clicked = await click_locator(primary_btn, timeout=5000, force=True)
    if not clicked:
        raise RuntimeError("Не удалось нажать основную кнопку сохранения")

    await page.wait_for_timeout(800)

    secondary_btn = side.locator("[data-test-component='undefined-footerSidePage-actionBarSecondaryButton']")
    await click_locator(secondary_btn, timeout=2000, force=True)
    await page.wait_for_timeout(300)


async def close_sidepage_if_open(page: Page) -> None:
    side = page.locator("[data-test-component='undefined-sidePage']")
    if await side.count() == 0:
        return

    try:
        visible = await side.first.is_visible()
    except Exception:
        visible = False

    if not visible:
        return

    secondary_btn = side.locator("[data-test-component='undefined-footerSidePage-actionBarSecondaryButton']")
    if await click_locator(secondary_btn, timeout=2000, force=True):
        await page.wait_for_timeout(300)
        return

    try:
        close_btn = side.locator("[data-test-component*='iconButton']").first
        if await click_locator(close_btn, timeout=2000, force=True):
            await page.wait_for_timeout(300)
            return
    except Exception:
        pass

    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)
    except Exception:
        pass


async def try_claim_student(
    student_name: str,
    processed_names: Set[str],
    in_progress_names: Set[str],
    state_lock: asyncio.Lock,
) -> bool:
    async with state_lock:
        if student_name in processed_names or student_name in in_progress_names:
            return False
        in_progress_names.add(student_name)
        return True


async def release_claim_student(
    student_name: str,
    in_progress_names: Set[str],
    state_lock: asyncio.Lock,
) -> None:
    async with state_lock:
        in_progress_names.discard(student_name)


async def wait_student_resolution(
    student_name: str,
    processed_names: Set[str],
    in_progress_names: Set[str],
    state_lock: asyncio.Lock,
    timeout_ms: int = 300_000,
    poll_ms: int = 500,
) -> str:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + timeout_ms / 1000

    while loop.time() < deadline:
        async with state_lock:
            if student_name in processed_names:
                return "processed"
            if student_name not in in_progress_names:
                return "free"
        await asyncio.sleep(poll_ms / 1000)

    return "timeout"


async def write_student_result(
    wb: Workbook,
    ws,
    agg: StudentAggregate,
    criteria: Dict[str, Any],
    status: str,
    answers: List[str],
    processed_names: Set[str],
    in_progress_names: Set[str],
    workbook_lock: asyncio.Lock,
    state_lock: asyncio.Lock,
) -> None:
    async with workbook_lock:
        append_student_row(ws, agg, criteria, status, answers)
        wb.save(XLSX_PATH)

    async with state_lock:
        if status == STATUS_SUCCESS:
            processed_names.add(agg.name)
        in_progress_names.discard(agg.name)


async def write_journal_summary(
    wb: Workbook,
    ws_summary,
    entry: JournalEntry,
    total_students: int,
    status_counts: Dict[str, int],
    workbook_lock: asyncio.Lock,
) -> None:
    async with workbook_lock:
        upsert_summary_row(
            ws_summary=ws_summary,
            entry=entry,
            total_students=total_students,
            status_counts=status_counts,
        )
        wb.save(XLSX_PATH)


def make_empty_status_counts() -> Dict[str, int]:
    return {status: 0 for status in ALL_STATUSES}


async def process_single_journal(
    context: BrowserContext,
    entry: JournalEntry,
    sem: asyncio.Semaphore,
    wb: Workbook,
    ws,
    ws_summary,
    processed_names: Set[str],
    in_progress_names: Set[str],
    state_lock: asyncio.Lock,
    workbook_lock: asyncio.Lock,
) -> None:
    async with sem:
        page = await context.new_page()
        journal_prefix = f"{entry.parallel} | {entry.class_name} | {entry.subject}"
        total_students = 0
        status_counts = make_empty_status_counts()

        try:
            print(f"[ЖУРНАЛ] Старт: {journal_prefix}")

            await safe_goto(page, entry.url)
            await page.wait_for_selector("table", timeout=REQUEST_TIMEOUT)

            print(f"[WAIT] {journal_prefix} | дополнительная стабилизация журнала: {JOURNAL_EXTRA_STABILIZATION_MS / 1000:.0f} сек")
            await page.wait_for_timeout(JOURNAL_EXTRA_STABILIZATION_MS)

            await ensure_all_periods_selected(page)

            raw_table = await extract_table_basic(page)
            popup_comment_map, lesson_comments = await collect_optional_comments(page, raw_table)

            local_students = build_local_students_from_table(
                entry=entry,
                raw_table=raw_table,
                popup_comment_map=popup_comment_map,
                journal_lesson_comments=lesson_comments,
            )

            total_students = len(local_students)
            print(f"[ЖУРНАЛ] {journal_prefix} | найдено учеников: {total_students}")

            for idx, local in enumerate(local_students, start=1):
                student_name = local.name

                claimed = await try_claim_student(
                    student_name=student_name,
                    processed_names=processed_names,
                    in_progress_names=in_progress_names,
                    state_lock=state_lock,
                )

                if not claimed:
                    resolution = await wait_student_resolution(
                        student_name=student_name,
                        processed_names=processed_names,
                        in_progress_names=in_progress_names,
                        state_lock=state_lock,
                    )

                    if resolution == "processed":
                        status_counts[STATUS_SUCCESS] += 1
                        print(f"  [SKIP->SUCCESS] {journal_prefix} | {student_name} уже успешно обработан ранее")
                        continue

                    if resolution == "free":
                        claimed = await try_claim_student(
                            student_name=student_name,
                            processed_names=processed_names,
                            in_progress_names=in_progress_names,
                            state_lock=state_lock,
                        )

                    if not claimed:
                        status_counts[STATUS_OTHER_ERROR] += 1
                        print(f"  [SKIP->ERROR] {journal_prefix} | {student_name} не удалось корректно захватить для обработки")
                        continue

                agg = agg_from_local(local)
                answers, criteria = recommend_answers_for_student(agg)
                answers_to_write = list(answers)

                print(
                    f"  [УЧЕНИК {idx}/{len(local_students)}] {journal_prefix} | {student_name} | "
                    f"avg_mark={criteria.get('avg_mark')} marks={criteria.get('mark_count')} n={criteria.get('n_count')}"
                )

                student_status = STATUS_OTHER_ERROR

                try:
                    if FILL_FORM_IN_UI:
                        status, q_meta = await fill_recommendation_in_ui_on_current_journal(page, student_name, answers)

                        if status == "already_assessed":
                            student_status = STATUS_SUCCESS
                            answers_to_write = EMPTY_ANSWERS
                            criteria["ui_status"] = "already_assessed"
                            criteria["note"] = "Ученик уже оценён ранее. Повторное заполнение не выполнялось."

                        elif status == "ouu_disabled":
                            student_status = STATUS_OUU_INACTIVE
                            criteria["ui_status"] = "ouu_button_inactive"
                            criteria["note"] = "Кнопка 'Оценить учебные умения' была неактивна."

                        elif status == "filled":
                            criteria["question_meta_from_ui"] = q_meta

                            if MANUAL_REVIEW_AFTER_FILL:
                                await wait_user_input(
                                    f"\nПроверьте ребёнка '{student_name}' "
                                    f"({entry.class_name} / {entry.subject}). "
                                    f"После проверки нажмите Enter..."
                                )

                            if AUTO_SAVE_AFTER_FILL:
                                try:
                                    await save_assessment_sidepage(page)
                                    student_status = STATUS_SUCCESS
                                    criteria["ui_status"] = "saved"
                                except SaveButtonInactiveError as e:
                                    student_status = STATUS_SAVE_INACTIVE
                                    criteria["ui_status"] = "save_button_inactive"
                                    criteria["save_error"] = str(e)
                                    await close_sidepage_if_open(page)
                            else:
                                await close_sidepage_if_open(page)
                                student_status = STATUS_OTHER_ERROR
                                criteria["ui_status"] = "not_saved_by_config"
                                criteria["note"] = "Автосохранение отключено конфигурацией, статус не может считаться успешным."
                    else:
                        student_status = STATUS_SUCCESS
                        criteria["ui_status"] = "ui_fill_disabled"

                except Exception as student_error:
                    student_status = STATUS_OTHER_ERROR
                    criteria["error"] = str(student_error)
                    await close_sidepage_if_open(page)

                try:
                    await write_student_result(
                        wb=wb,
                        ws=ws,
                        agg=agg,
                        criteria=criteria,
                        status=student_status,
                        answers=answers_to_write,
                        processed_names=processed_names,
                        in_progress_names=in_progress_names,
                        workbook_lock=workbook_lock,
                        state_lock=state_lock,
                    )
                    status_counts[student_status] += 1
                    print(f"    [XLSX] {student_name} | статус: {student_status}")
                except Exception as write_error:
                    await release_claim_student(
                        student_name=student_name,
                        in_progress_names=in_progress_names,
                        state_lock=state_lock,
                    )
                    status_counts[STATUS_OTHER_ERROR] += 1
                    print(f"    [ERROR] {journal_prefix} | {student_name} | ошибка записи XLSX: {write_error}")

            print(f"[ЖУРНАЛ] Готово: {journal_prefix}")

        except Exception as e:
            print(f"[ОШИБКА] Журнал {journal_prefix} | {e}")

        finally:
            try:
                await write_journal_summary(
                    wb=wb,
                    ws_summary=ws_summary,
                    entry=entry,
                    total_students=total_students,
                    status_counts=status_counts,
                    workbook_lock=workbook_lock,
                )
                print(
                    f"[СВОДКА] {journal_prefix} | total={total_students}, "
                    f"success={status_counts[STATUS_SUCCESS]}, "
                    f"ouu_inactive={status_counts[STATUS_OUU_INACTIVE]}, "
                    f"save_inactive={status_counts[STATUS_SAVE_INACTIVE]}, "
                    f"other_error={status_counts[STATUS_OTHER_ERROR]}"
                )
            except Exception as summary_error:
                print(f"[ОШИБКА] Не удалось записать сводку по журналу {journal_prefix} | {summary_error}")

            try:
                await page.close()
            except Exception:
                pass


async def main() -> None:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    print(f"[CONFIG] Yandex Browser: {YANDEX_BROWSER_PATH}")
    print(f"[CONFIG] User Data: {USER_DATA_DIR}")
    print(f"[CONFIG] Profile: {PROFILE_DIRECTORY}")
    print(f"[CONFIG] XLSX: {XLSX_PATH}")

    wb, ws, ws_summary = ensure_workbook(XLSX_PATH)
    processed_names = load_processed_names(ws)
    summary_map = load_summary_completion_map(ws_summary)
    in_progress_names: Set[str] = set()

    state_lock = asyncio.Lock()
    workbook_lock = asyncio.Lock()

    playwright, context, page = await init_browser(headless=HEADLESS)

    try:
        print("[1/3] Открываю список журналов...")
        entries = await scrape_journal_entries(page)
        print(f"[INFO] Найдено журналов: {len(entries)}")
        print(f"[INFO] Параллельных вкладок: {MAX_CONCURRENT_JOURNALS}")

        if not entries:
            print("[STOP] Журналы не найдены.")
            return

        filtered_entries: List[JournalEntry] = []
        skipped_by_summary = 0

        for entry in entries:
            stats = summary_map.get(entry.url)
            if stats and stats.get("total", 0) > 0 and stats.get("success", 0) >= stats.get("total", 0):
                skipped_by_summary += 1
                print(
                    f"[SKIP] Журнал уже полностью завершён по сводке: "
                    f"{entry.parallel} | {entry.class_name} | {entry.subject} | {entry.url}"
                )
                continue
            filtered_entries.append(entry)

        print(f"[INFO] Пропущено по сводке: {skipped_by_summary}")
        print(f"[INFO] К обработке осталось журналов: {len(filtered_entries)}")

        if not filtered_entries:
            print(f"[3/3] Готово. Все журналы уже завершены по сводной таблице: {XLSX_PATH}")
            return

        print("[2/3] Запускаю параллельную обработку журналов...")

        sem = asyncio.Semaphore(MAX_CONCURRENT_JOURNALS)
        tasks = [
            process_single_journal(
                context=context,
                entry=entry,
                sem=sem,
                wb=wb,
                ws=ws,
                ws_summary=ws_summary,
                processed_names=processed_names,
                in_progress_names=in_progress_names,
                state_lock=state_lock,
                workbook_lock=workbook_lock,
            )
            for entry in filtered_entries
        ]

        results = await asyncio.gather(*tasks, return_exceptions=True)
        for result in results:
            if isinstance(result, Exception):
                print(f"[ОШИБКА] Задача завершилась с ошибкой: {result}")

        print(f"[3/3] Готово. Таблица сохранена: {XLSX_PATH}")

    finally:
        try:
            await context.close()
        except Exception:
            pass
        try:
            await playwright.stop()
        except Exception:
            pass


def run_application() -> None:
    config = show_startup_dialog(APP_CONFIG)
    if config is None:
        print("[STOP] Запуск отменён пользователем.")
        return

    apply_config(config)
    asyncio.run(main())


if __name__ == "__main__":
    run_application()